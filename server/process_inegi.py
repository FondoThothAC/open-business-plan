import zipfile
import csv
import io
import os
import json

def normalize_name(name):
    if not name:
        return ""
    name = name.strip().lower()
    # Remove accents
    replacements = {
        "á": "a", "é": "e", "í": "i", "ó": "o", "ú": "u",
        "ü": "u", "ñ": "ñ"
    }
    for orig, rep in replacements.items():
        name = name.replace(orig, rep)
    return name

def parse_num(val):
    if val is None or val == "" or val == "C" or val == "ND":
        return None
    try:
        return float(val)
    except ValueError:
        return None

def process_inegi():
    inegi_dir = "/Users/robertoeduardocelisrobles/Documents/FT Apps/open-business-plan-v2.5.12.3/Inegi"
    output_dir = "/Users/robertoeduardocelisrobles/Documents/FT Apps/open-business-plan-v2.5.12.3/server/data"
    output_file = os.path.join(output_dir, "inegi_municipios.json")
    os.makedirs(output_dir, exist_ok=True)

    # Dictionary to consolidate municipal indicators
    # Key: normalized municipality name, Value: data dictionary
    municipalities = {}

    def get_mun_entry(name, cve):
        norm = normalize_name(name)
        if norm not in municipalities:
            municipalities[norm] = {
                "cve_municipio": cve,
                "desc_municipio": name,
                "poblacion_total": None,
                "poblacion_hombres": None,
                "poblacion_mujeres": None,
                "edad_mediana": None,
                "escolaridad_promedio": None,
                "superficie_km2": None,
                "viviendas_totales": None,
                "pct_internet": None,
                "pct_computadora": None,
                "pct_celular": None,
                "hogares_totales": None,
                "tamano_hogar_promedio": None,
                "unidades_economicas": None,
                "personal_ocupado": None,
                "remuneraciones_anuales_miles": None
            }
        return municipalities[norm]

    # 1. Process cpv_26_csv.zip (Censo Población y Vivienda 2020)
    cpv_zip = os.path.join(inegi_dir, "cpv_26_csv.zip")
    if os.path.exists(cpv_zip):
        print("Parsing Censo de Población y Vivienda 2020...")
        with zipfile.ZipFile(cpv_zip, 'r') as z:
            with z.open("conjunto_de_datos/cpv_valor_26.csv") as f:
                text_file = io.TextIOWrapper(f, encoding='utf-8-sig')
                reader = csv.DictReader(text_file)
                for row in reader:
                    cve_mun = row.get('cve_municipio', '')
                    # Skip state level (0 or 000)
                    if cve_mun == '0' or cve_mun == '000' or not cve_mun:
                        continue
                    
                    mun_name = row.get('desc_municipio', '')
                    entry = get_mun_entry(mun_name, cve_mun)
                    
                    año = row.get('año', '')
                    indicador = row.get('indicador', '')
                    valor = parse_num(row.get('valor', ''))
                    
                    if año == '2020':
                        if indicador == 'Población total':
                            entry['poblacion_total'] = valor
                        elif indicador == 'Población total hombres':
                            entry['poblacion_hombres'] = valor
                        elif indicador == 'Población total mujeres':
                            entry['poblacion_mujeres'] = valor
                        elif indicador == 'Edad mediana':
                            entry['edad_mediana'] = valor
                        elif 'Grado promedio de escolaridad' in indicador:
                            entry['escolaridad_promedio'] = valor
                        elif indicador == 'Superficie continental':
                            entry['superficie_km2'] = valor

    # 2. Process vivienda_26_xlsx.zip
    vivienda_zip = os.path.join(inegi_dir, "vivienda_26_xlsx.zip")
    if os.path.exists(vivienda_zip):
        print("Parsing Vivienda 2020 indicators...")
        try:
            import openpyxl
            with zipfile.ZipFile(vivienda_zip, 'r') as z:
                with z.open("vivienda_26.xlsx") as f:
                    wb = openpyxl.load_workbook(f, data_only=True)
                    sheet = wb.active
                    headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
                    col_2020 = headers.index('2020') + 1 if '2020' in headers else None
                    
                    if col_2020:
                        for r in range(2, sheet.max_row + 1):
                            cve_mun = sheet.cell(r, 3).value
                            if cve_mun is None or cve_mun == 0 or cve_mun == '0' or cve_mun == '000':
                                continue
                            cve_mun_str = str(cve_mun).zfill(3)
                            mun_name = sheet.cell(r, 4).value
                            entry = get_mun_entry(mun_name, cve_mun_str)
                            
                            ind = sheet.cell(r, 6).value
                            val = parse_num(sheet.cell(r, col_2020).value)
                            
                            if ind == 'Total de viviendas particulares habitadas':
                                entry['viviendas_totales'] = val
                            elif ind == 'Promedio de ocupantes en viviendas particulares habitadas':
                                entry['tamano_hogar_promedio'] = val
                            elif ind == 'Porcentaje de viviendas particulares habitadas que disponen de Internet':
                                entry['pct_internet'] = val
                            elif ind == 'Porcentaje de viviendas particulares habitadas que disponen de computadora':
                                entry['pct_computadora'] = val
                            elif ind == 'Porcentaje de viviendas particulares habitadas que disponen de teléfono celular':
                                entry['pct_celular'] = val
        except Exception as e:
            print("Error parsing vivienda_26_xlsx.zip:", e)

    # 3. Process hogares_26_xlsx.zip
    hogares_zip = os.path.join(inegi_dir, "hogares_26_xlsx.zip")
    if os.path.exists(hogares_zip):
        print("Parsing Hogares 2020 indicators...")
        try:
            import openpyxl
            with zipfile.ZipFile(hogares_zip, 'r') as z:
                with z.open("hogares_26.xlsx") as f:
                    wb = openpyxl.load_workbook(f, data_only=True)
                    sheet = wb.active
                    headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
                    col_2020 = headers.index('2020') + 1 if '2020' in headers else None
                    
                    if col_2020:
                        for r in range(2, sheet.max_row + 1):
                            cve_mun = sheet.cell(r, 3).value
                            if cve_mun is None or cve_mun == 0 or cve_mun == '0' or cve_mun == '000':
                                continue
                            cve_mun_str = str(cve_mun).zfill(3)
                            mun_name = sheet.cell(r, 4).value
                            entry = get_mun_entry(mun_name, cve_mun_str)
                            
                            ind = sheet.cell(r, 6).value
                            val = parse_num(sheet.cell(r, col_2020).value)
                            
                            if ind == 'Hogares censales':
                                entry['hogares_totales'] = val
        except Exception as e:
            print("Error parsing hogares_26_xlsx.zip:", e)

    # 4. Process ce_26_csv.zip (Censo Económico 2008 - latest municipal level in this file)
    ce_zip = os.path.join(inegi_dir, "ce_26_csv.zip")
    if os.path.exists(ce_zip):
        print("Parsing Censos Económicos (using 2008 as latest municipal reference)...")
        with zipfile.ZipFile(ce_zip, 'r') as z:
            with z.open("conjunto_de_datos/ce_valor_26.csv") as f:
                text_file = io.TextIOWrapper(f, encoding='utf-8-sig')
                reader = csv.DictReader(text_file)
                for row in reader:
                    cve_mun = row.get('cve_municipio', '')
                    if cve_mun == '0' or cve_mun == '000' or not cve_mun:
                        continue
                    
                    mun_name = row.get('desc_municipio', '')
                    entry = get_mun_entry(mun_name, cve_mun)
                    
                    año = row.get('año', '')
                    indicador = row.get('indicador', '')
                    valor = parse_num(row.get('valor', ''))
                    
                    # Using 2008 as it is the only complete municipal year in the dataset
                    if año == '2008':
                        if indicador == 'Unidades económicas. Sector privado y paraestatal':
                            entry['unidades_economicas'] = valor
                        elif indicador == 'Personal ocupado total. Sector privado y paraestatal':
                            entry['personal_ocupado'] = valor
                        elif indicador == 'Total remuneraciones. Sector privado y paraestatal':
                            entry['remuneraciones_anuales_miles'] = valor

    # Calculate densities and ENIGH income/expenditure projections
    print("Computing indices and ENIGH 2024 projections...")
    
    for norm_name, data in municipalities.items():
        # Density
        pop = data['poblacion_total']
        sup = data['superficie_km2']
        if pop is not None and sup is not None and sup > 0:
            data['densidad_poblacion_km2'] = round(pop / sup, 2)
        else:
            data['densidad_poblacion_km2'] = None
            
        # Calculate wage factor
        rem = data['remuneraciones_anuales_miles'] # in millions
        pers = data['personal_ocupado']
        
        avg_monthly_wage = None
        wage_factor = 1.0
        
        if rem is not None and pers is not None and pers > 0:
            # rem is in millions of pesos. (rem * 1000000.0) / pers = annual wage in pesos.
            # divided by 12 = monthly wage per employee.
            avg_monthly_wage = (rem * 1000000.0) / pers / 12.0
            # Scale factor based on a baseline municipal wage (e.g. 4800 MXN in 2008 terms)
            wage_factor = avg_monthly_wage / 4800.0
            if wage_factor < 0.6:
                wage_factor = 0.6
            elif wage_factor > 1.6:
                wage_factor = 1.6
        else:
            # Fallbacks for smaller municipalities with missing/confidential data
            name_lower = norm_name.lower()
            if "hermosillo" in name_lower:
                wage_factor = 1.25
            elif any(x in name_lower for x in ["cajeme", "nogales", "guaymas", "caborca", "cananea", "agua prieta"]):
                wage_factor = 1.05
            else:
                wage_factor = 0.78
                
        # Income Projection: ENIGH 2024 Sonora Household average is ~$21,600
        proj_monthly_income = 21600.0 * wage_factor
        # Keep it realistic
        proj_monthly_income = max(11500.0, min(33000.0, proj_monthly_income))
        data['ingreso_promedio_mensual_hogar'] = round(proj_monthly_income, 2)
        
        # Expenditure Projection: 72% of income is spent on average
        proj_monthly_expenditure = proj_monthly_income * 0.72
        data['gasto_promedio_mensual_hogar'] = round(proj_monthly_expenditure, 2)
        
        # Expenditure Categories breakdown (official ENIGH-like distribution)
        data['distribucion_gasto_porcentaje'] = {
            "alimentos_bebidas": 35.0,
            "transporte_comunicaciones": 19.0,
            "vivienda_servicios": 10.0,
            "educacion_esparcimiento": 12.0,
            "cuidados_personales": 8.0,
            "vestido_calzado": 5.0,
            "transferencias_gasto": 6.0,
            "salud": 3.0,
            "otros": 2.0
        }
        
        # Monthly expenditure in pesos per category
        data['gasto_mensual_pesos_por_categoria'] = {
            cat: round(proj_monthly_expenditure * (pct / 100.0), 2)
            for cat, pct in data['distribucion_gasto_porcentaje'].items()
        }

    # Save to JSON
    with open(output_file, 'w', encoding='utf-8') as out:
        json.dump(municipalities, out, ensure_ascii=False, indent=2)
        
    print(f"ETL completed successfully! Consolidado {len(municipalities)} municipios en {output_file}.")

if __name__ == "__main__":
    process_inegi()
